import asyncio
import json
from datetime import timedelta

import pytest
from fastapi import HTTPException
from sqlalchemy import inspect as sa_inspect
from sqlmodel import Session, SQLModel, create_engine, select

from app.models.db import Conversation, Message, PendingToolAction, Project, ProjectFile, ProjectMember, User
from app.routers import chat_actions
from app.services.chat.action_project_writes import (
    cleanup_prepared_project_write,
    persist_prepared_project_write,
)
from app.services.chat.action_reaper import reap_stale_executing_actions
from app.services.project_documents import create_markdown_project_file
from app.services.time_utils import utc_now_naive
from app.tools import office_documents, project_markdown


def _setup_action(tmp_path):
    bind = create_engine(
        f"sqlite:///{tmp_path / 'hitas-final-auth.sqlite'}",
        connect_args={"check_same_thread": False},
    )
    SQLModel.metadata.create_all(bind)
    uploads = tmp_path / "uploads"
    (uploads / "generated").mkdir(parents=True)
    with Session(bind) as session:
        project = Project(name="Final auth", client="Client")
        actor = User(email="writer@example.com", password_hash="x")
        session.add(project)
        session.add(actor)
        session.commit()
        session.refresh(project)
        session.refresh(actor)
        member = ProjectMember(project_id=project.id, user_id=actor.id, role="editor")
        conversation = Conversation(title="Approval", project_id=project.id)
        session.add(member)
        session.add(conversation)
        session.commit()
        session.refresh(conversation)
        tool_input = {
            "project_id": project.id,
            "file_type": "pdf",
            "file_name": "deliverable.pdf",
            "title": "Deliverable",
            "content": "Approved content",
        }
        action = PendingToolAction(
            conversation_id=conversation.id,
            project_id=project.id,
            tool_name=office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input_json=json.dumps(tool_input),
            action_type="write_office_document",
            title="Create deliverable",
            status="executing",
            confirmed_by_user_id=actor.id,
            confirmed_at=utc_now_naive(),
        )
        session.add(action)
        session.commit()
        session.refresh(action)
        return bind, uploads, int(project.id), int(actor.id), int(member.id), int(action.id), tool_input


def _prepared_pdf(uploads, name="prepared.pdf"):
    path = uploads / "generated" / name
    path.write_bytes(b"prepared-pdf")
    return {
        "kind": "office_create",
        "cleanup_source": True,
        "source_path": str(path),
        "file_name": "deliverable.pdf",
        "file_type": "pdf",
        "folder_id": None,
        "summary": "approved",
        "preview_text": "approved",
    }


@pytest.mark.parametrize("revocation", ["inactive", "removed", "viewer"])
def test_office_finalization_discards_prepared_file_after_actor_revocation(
    monkeypatch,
    tmp_path,
    revocation,
):
    bind, uploads, project_id, actor_id, member_id, action_id, tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    prepared_path = uploads / "generated" / f"{revocation}.pdf"

    async def fake_prepare(_bind, _tool_name, _tool_input):
        prepared = _prepared_pdf(uploads, prepared_path.name)
        with Session(bind) as session:
            if revocation == "inactive":
                actor = session.get(User, actor_id)
                actor.is_active = False
                session.add(actor)
            elif revocation == "removed":
                member = session.get(ProjectMember, member_id)
                session.delete(member)
            else:
                member = session.get(ProjectMember, member_id)
                member.role = "viewer"
                session.add(member)
            session.commit()
        return prepared

    monkeypatch.setattr(chat_actions, "prepare_pending_project_write", fake_prepare)
    finalized = asyncio.run(
        chat_actions._execute_final_authorized_project_write(
            bind,
            action_id,
            office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input,
            emit_message=True,
        )
    )

    assert finalized["status"] == "failed"
    assert not prepared_path.exists()
    with Session(bind) as session:
        assert session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all() == []
        messages = session.exec(select(Message)).all()
        assert len(messages) == 1
        assert "执行失败" in messages[0].content
        action = session.get(PendingToolAction, action_id)
        assert action.status == "failed"
        assert action.error_message
        assert json.loads(action.result_json)["success"] is False


@pytest.mark.parametrize("terminal_status", ["rejected", "superseded"])
def test_office_finalization_does_not_overwrite_canceled_or_superseded_action(
    monkeypatch,
    tmp_path,
    terminal_status,
):
    bind, uploads, project_id, _actor_id, _member_id, action_id, tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    prepared_path = uploads / "generated" / f"{terminal_status}.pdf"

    async def fake_prepare(_bind, _tool_name, _tool_input):
        prepared = _prepared_pdf(uploads, prepared_path.name)
        with Session(bind) as session:
            action = session.get(PendingToolAction, action_id)
            action.status = terminal_status
            action.error_message = "canceled elsewhere"
            session.add(action)
            session.commit()
        return prepared

    monkeypatch.setattr(chat_actions, "prepare_pending_project_write", fake_prepare)
    finalized = asyncio.run(
        chat_actions._execute_final_authorized_project_write(
            bind,
            action_id,
            office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input,
            emit_message=True,
        )
    )

    assert finalized["status"] == terminal_status
    assert not prepared_path.exists()
    with Session(bind) as session:
        assert session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all() == []
        assert session.exec(select(Message)).all() == []
        assert session.get(PendingToolAction, action_id).status == terminal_status


def test_office_finalization_does_not_add_receipt_after_reaper_wins(monkeypatch, tmp_path):
    bind, uploads, project_id, _actor_id, _member_id, action_id, tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    prepared_path = uploads / "generated" / "reaped.pdf"

    async def fake_prepare(_bind, _tool_name, _tool_input):
        prepared = _prepared_pdf(uploads, prepared_path.name)
        with Session(bind) as session:
            action = session.get(PendingToolAction, action_id)
            action.confirmed_at = utc_now_naive() - timedelta(hours=2)
            session.add(action)
            session.commit()
            assert reap_stale_executing_actions(session, stale_after_minutes=30) == 1
        return prepared

    monkeypatch.setattr(chat_actions, "prepare_pending_project_write", fake_prepare)
    finalized = asyncio.run(
        chat_actions._execute_final_authorized_project_write(
            bind,
            action_id,
            office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input,
            emit_message=True,
        )
    )

    assert finalized["status"] == "failed"
    assert not prepared_path.exists()
    with Session(bind) as session:
        assert session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all() == []
        messages = session.exec(select(Message)).all()
        assert len(messages) == 1
        assert "人工核查" in messages[0].content


def test_office_success_commits_file_action_and_receipt_together(monkeypatch, tmp_path):
    bind, uploads, project_id, _actor_id, _member_id, action_id, tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    prepared_path = uploads / "generated" / "success.pdf"

    async def fake_prepare(_bind, _tool_name, _tool_input):
        return _prepared_pdf(uploads, prepared_path.name)

    monkeypatch.setattr(chat_actions, "prepare_pending_project_write", fake_prepare)
    original_payload_builder = chat_actions._action_finalization_payload
    payload_expired_states = []

    def observe_payload_state(action, **kwargs):
        payload_expired_states.append(sa_inspect(action).expired)
        return original_payload_builder(action, **kwargs)

    monkeypatch.setattr(chat_actions, "_action_finalization_payload", observe_payload_state)
    finalized = asyncio.run(
        chat_actions._execute_final_authorized_project_write(
            bind,
            action_id,
            office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input,
            emit_message=True,
        )
    )

    assert finalized["status"] == "completed"
    assert finalized["message_id"] is not None
    assert payload_expired_states == [False]
    assert not prepared_path.exists()
    with Session(bind) as session:
        files = session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all()
        assert len(files) == 1
        assert (uploads / files[0].path).read_bytes() == b"prepared-pdf"
        action = session.get(PendingToolAction, action_id)
        assert action.status == "completed"
        assert json.loads(action.result_json)["project_file_id"] == files[0].id
        messages = session.exec(select(Message)).all()
        assert len(messages) == 1
        assert "已执行" in messages[0].content


def test_markdown_action_uses_same_final_authorized_transaction(monkeypatch, tmp_path):
    bind, uploads, project_id, _actor_id, _member_id, action_id, _tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    tool_input = {
        "project_id": project_id,
        "mode": "create",
        "file_name": "decision-log.md",
        "content": "# Decision log\n\nApproved.",
    }
    with Session(bind) as session:
        action = session.get(PendingToolAction, action_id)
        action.tool_name = project_markdown.PROJECT_MARKDOWN_TOOL_NAME
        action.tool_input_json = json.dumps(tool_input)
        session.add(action)
        session.commit()

    finalized = asyncio.run(
        chat_actions._execute_final_authorized_project_write(
            bind,
            action_id,
            project_markdown.PROJECT_MARKDOWN_TOOL_NAME,
            tool_input,
            emit_message=True,
        )
    )

    assert finalized["status"] == "completed"
    with Session(bind) as session:
        project_file = session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).one()
        assert project_file.name == "decision-log.md"
        assert (uploads / project_file.path).read_text(encoding="utf-8") == tool_input["content"]
        assert session.get(PendingToolAction, action_id).status == "completed"
        assert len(session.exec(select(Message)).all()) == 1


def test_office_edit_prepares_copy_then_reauthorizes_before_overwrite(monkeypatch, tmp_path):
    from openpyxl import Workbook, load_workbook

    bind, uploads, project_id, _actor_id, _member_id, action_id, _tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    project_dir = uploads / "projects" / str(project_id)
    project_dir.mkdir(parents=True)
    source_path = project_dir / "source.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Sheet1"
    worksheet["A1"] = "before"
    workbook.save(source_path)

    with Session(bind) as session:
        project_file = ProjectFile(
            project_id=project_id,
            name="source.xlsx",
            file_type="xlsx",
            path=str(source_path.relative_to(uploads)),
            size_bytes=source_path.stat().st_size,
        )
        session.add(project_file)
        session.commit()
        session.refresh(project_file)
        tool_input = {
            "project_id": project_id,
            "file_id": project_file.id,
            "edits": [{"action": "update_cell", "sheet": "Sheet1", "cell": "A1", "value": "after"}],
        }
        action = session.get(PendingToolAction, action_id)
        action.tool_name = office_documents.EDIT_PROJECT_OFFICE_DOCUMENT_TOOL_NAME
        action.tool_input_json = json.dumps(tool_input)
        session.add(action)
        session.commit()

    finalized = asyncio.run(
        chat_actions._execute_final_authorized_project_write(
            bind,
            action_id,
            office_documents.EDIT_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input,
            emit_message=True,
        )
    )

    assert finalized["status"] == "completed"
    edited = load_workbook(source_path)
    assert edited["Sheet1"]["A1"].value == "after"
    with Session(bind) as session:
        assert session.get(PendingToolAction, action_id).status == "completed"
        assert len(session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all()) == 1
        assert len(session.exec(select(Message)).all()) == 1
    assert not list((uploads / "generated").glob("hitas_edit_*"))


def test_final_authorization_validation_failure_terminalizes_exact_generation(
    monkeypatch,
    tmp_path,
):
    bind, uploads, project_id, _actor_id, _member_id, action_id, tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    prepared_path = uploads / "generated" / "signature-failure.pdf"

    async def fake_prepare(_bind, _tool_name, _tool_input):
        return _prepared_pdf(uploads, prepared_path.name)

    original_validate = chat_actions._validated_execution_tool_input
    validation_count = 0

    def validate_then_fail(session, action):
        nonlocal validation_count
        validation_count += 1
        if validation_count == 1:
            return original_validate(session, action)
        raise HTTPException(409, "Approval signature changed after prepare")

    monkeypatch.setattr(chat_actions, "prepare_pending_project_write", fake_prepare)
    monkeypatch.setattr(
        chat_actions,
        "_validated_execution_tool_input",
        validate_then_fail,
    )

    finalized = asyncio.run(
        chat_actions._execute_final_authorized_project_write(
            bind,
            action_id,
            office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input,
            emit_message=True,
        )
    )

    assert finalized["status"] == "failed"
    assert "signature changed" in finalized["error_message"]
    assert not prepared_path.exists()
    with Session(bind) as session:
        action = session.get(PendingToolAction, action_id)
        assert action.status == "failed"
        assert session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all() == []
        assert len(session.exec(select(Message)).all()) == 1


def test_final_authorization_terminalizer_preserves_changed_executing_generation(
    monkeypatch,
    tmp_path,
):
    bind, uploads, project_id, _actor_id, _member_id, action_id, tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    prepared_path = uploads / "generated" / "changed-generation.pdf"
    replacement_input = {**tool_input, "file_name": "replacement.pdf"}

    async def fake_prepare(_bind, _tool_name, _tool_input):
        prepared = _prepared_pdf(uploads, prepared_path.name)
        with Session(bind) as session:
            action = session.get(PendingToolAction, action_id)
            action.tool_input_json = json.dumps(replacement_input)
            session.add(action)
            session.commit()
        return prepared

    monkeypatch.setattr(chat_actions, "prepare_pending_project_write", fake_prepare)
    finalized = asyncio.run(
        chat_actions._execute_final_authorized_project_write(
            bind,
            action_id,
            office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input,
            emit_message=True,
        )
    )

    assert finalized["status"] == "executing"
    assert finalized["suppress_followup_receipt"] is True
    assert not prepared_path.exists()
    with Session(bind) as session:
        action = session.get(PendingToolAction, action_id)
        assert action.status == "executing"
        assert json.loads(action.tool_input_json) == replacement_input
        assert session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all() == []
        assert session.exec(select(Message)).all() == []


def test_final_authorization_batch_terminalizes_failure_and_skips_remainder(
    monkeypatch,
    tmp_path,
):
    bind, uploads, project_id, actor_id, _member_id, action_id, tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)
    batch_id = "final-auth-terminal-batch"
    with Session(bind) as session:
        first = session.get(PendingToolAction, action_id)
        first.approval_batch_id = batch_id
        first.sequence_index = 0
        second = PendingToolAction(
            conversation_id=first.conversation_id,
            project_id=project_id,
            tool_name=office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input_json=json.dumps({**tool_input, "file_name": "second.pdf"}),
            action_type=first.action_type,
            title="Create second deliverable",
            status="executing",
            approval_batch_id=batch_id,
            sequence_index=1,
            confirmed_by_user_id=actor_id,
            confirmed_at=first.confirmed_at,
        )
        session.add(first)
        session.add(second)
        session.commit()
        session.refresh(second)
        second_id = int(second.id or 0)
        conversation_id = int(first.conversation_id)

    prepare_calls = 0

    async def prepare_then_revoke(_bind, _tool_name, _tool_input):
        nonlocal prepare_calls
        prepare_calls += 1
        prepared = _prepared_pdf(uploads, "batch-first.pdf")
        with Session(bind) as session:
            actor = session.get(User, actor_id)
            actor.is_active = False
            session.add(actor)
            session.commit()
        return prepared

    monkeypatch.setattr(chat_actions, "prepare_pending_project_write", prepare_then_revoke)
    response = asyncio.run(
        chat_actions._execute_batch_actions(
            bind,
            batch_id,
            [
                {
                    "id": action_id,
                    "tool_name": office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
                    "tool_input": tool_input,
                    "conversation_id": conversation_id,
                    "project_id": project_id,
                    "confirmed_by_user_id": actor_id,
                    "recovery_guarded": False,
                },
                {
                    "id": second_id,
                    "tool_name": office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
                    "tool_input": {**tool_input, "file_name": "second.pdf"},
                    "conversation_id": conversation_id,
                    "project_id": project_id,
                    "confirmed_by_user_id": actor_id,
                    "recovery_guarded": False,
                },
            ],
        )
    )

    assert response.status == "failed"
    assert prepare_calls == 1
    with Session(bind) as session:
        actions = session.exec(
            select(PendingToolAction)
            .where(PendingToolAction.approval_batch_id == batch_id)
            .order_by(PendingToolAction.sequence_index)
        ).all()
        assert [action.status for action in actions] == ["failed", "skipped"]
        assert session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all() == []
        assert len(session.exec(select(Message)).all()) == 1


def test_project_conversation_global_tool_is_not_blocked_as_project_registry_write(tmp_path):
    bind, _uploads, _project_id, _actor_id, _member_id, action_id, _tool_input = _setup_action(tmp_path)
    global_input = {"file_name": "notes.txt", "content": "approved"}
    with Session(bind) as session:
        action = session.get(PendingToolAction, action_id)
        action.tool_name = "save_text"
        action.tool_input_json = json.dumps(global_input)
        session.add(action)
        session.commit()

    assert (
        chat_actions._atomic_write_boundary_error_before_execution(
            bind,
            action_id=action_id,
            expected_tool_name="save_text",
            expected_tool_input=global_input,
        )
        == ""
    )


def test_background_final_authorization_failure_terminalizes_exact_generation(
    monkeypatch,
    tmp_path,
):
    bind, uploads, project_id, actor_id, _member_id, action_id, tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    monkeypatch.setattr(project_markdown, "UPLOADS_DIR", uploads)

    async def prepare_then_deactivate(_bind, _tool_name, _tool_input):
        prepared = _prepared_pdf(uploads, "background-revoked.pdf")
        with Session(bind) as session:
            actor = session.get(User, actor_id)
            actor.is_active = False
            session.add(actor)
            session.commit()
        return prepared

    monkeypatch.setattr(chat_actions, "prepare_pending_project_write", prepare_then_deactivate)
    asyncio.run(
        chat_actions._execute_action_in_background(
            bind,
            action_id,
            office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
            tool_input,
        )
    )

    with Session(bind) as session:
        action = session.get(PendingToolAction, action_id)
        assert action.status == "failed"
        assert action.error_message
        assert session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all() == []
        assert len(session.exec(select(Message)).all()) == 1


def test_outer_transaction_failure_removes_uncommitted_office_copy(monkeypatch, tmp_path):
    bind, uploads, project_id, _actor_id, _member_id, _action_id, _tool_input = _setup_action(tmp_path)
    monkeypatch.setattr(office_documents, "UPLOADS_DIR", uploads)
    prepared = _prepared_pdf(uploads, "rollback.pdf")

    with Session(bind) as session:
        project = session.get(Project, project_id)
        with pytest.raises(RuntimeError, match="force outer rollback"):
            with persist_prepared_project_write(
                session,
                project=project,
                tool_name=office_documents.WRITE_PROJECT_OFFICE_DOCUMENT_TOOL_NAME,
                prepared=prepared,
            ):
                raise RuntimeError("force outer rollback")
        session.rollback()

    with Session(bind) as session:
        assert session.exec(select(ProjectFile).where(ProjectFile.project_id == project_id)).all() == []
    project_dir = uploads / "projects" / str(project_id)
    assert not project_dir.exists() or not [path for path in project_dir.iterdir() if path.is_file()]
    cleanup_prepared_project_write(prepared)


def test_markdown_flush_failure_removes_orphan_file(monkeypatch, tmp_path):
    bind = create_engine(f"sqlite:///{tmp_path / 'markdown-orphan.sqlite'}")
    SQLModel.metadata.create_all(bind)
    uploads = tmp_path / "uploads"
    with Session(bind) as session:
        project = Project(name="Markdown", client="Client")
        session.add(project)
        session.commit()
        session.refresh(project)
        project_dir = uploads / "projects" / str(project.id)

        def fail_flush(*_args, **_kwargs):
            raise RuntimeError("flush failed")

        monkeypatch.setattr(session, "flush", fail_flush)
        with pytest.raises(RuntimeError, match="flush failed"):
            create_markdown_project_file(
                session,
                int(project.id),
                "orphan.md",
                "# content",
                uploads_dir=uploads,
                commit=False,
            )

        assert project_dir.exists()
        assert list(project_dir.iterdir()) == []
